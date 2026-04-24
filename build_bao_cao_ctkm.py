"""
Script build Báo cáo Rà soát CTKM đang hoạt động theo tháng.

INPUT (đặt cùng thư mục):
  1. data campaign dang hoat dong tN.xlsx   (export từ tool, 5 cột: Campaign | StartDate | EndDate | Salon | Đã sử dụng; header ở dòng 3, data từ dòng 4)
  2. Chia salon cụm Salesup - phân cụm mới tháng N.csv   (danh sách salon công ty ở cột 2)

OUTPUT:
  bao-cao-ctkm-dang-hoat-dong-tN.html

CÁCH CHẠY:
  python build_bao_cao_ctkm.py --month 4 --today 2026-04-24
  (Mặc định: tháng hiện tại và ngày hôm nay)

QUY TẮC THỐNG KÊ (đã được user duyệt qua các vòng feedback T4/2026):
  - Lọc salon: chỉ giữ camp có phát sinh ở Salon Công ty (58 salon trong CSV) + "Shop".
    Bỏ toàn bộ NQ (28 NK, 99 LVV, 151 CG, 1109 SHGL...) và các camp không còn salon công ty sau lọc.
  - Bỏ category MOYO hoàn toàn (toàn bộ Thẻ Chị đẹp MOYO, MOYO member, Tôi yêu 30Shine - Moyo, 28 NK MOYO...)
  - Bỏ camp đã hết hạn (end_date < today).
  - Sort trong mỗi nhóm: days_left ASC, rồi -total_used DESC (camp sắp hết lên đầu, camp dùng nhiều lên trước).

PHÂN NHÓM (7 category, theo thứ tự hiển thị):
  1. T4       — CTKM Trọng tâm Tháng 4 (camp marketing chủ lực cho tháng)
  2. SP       — Khuyến mãi Sản phẩm (Laborie combo, Vu Lan máy massage...)
  3. ECOM     — CTKM Ecom / Online (prefix "Ecom -" hoặc word boundary \becom\b)
  4. MEMBER   — Ưu đãi Shine Member & Thẻ Dịch vụ (gồm cả Tôi yêu 30Shine, Shinecombo BHXH, thẻ N buổi HSD...)
  5. VOUCHER  — Voucher Code / MKT (gồm #2909 [CAMPAIGN KHÔNG GIA HẠN] thẻ giảm 50k/3 lần)
  6. KHAC     — Ưu đãi khác / Salon riêng (còn lại)
  7. NOIBO    — Camp Nội bộ / Sự cố (VH_sự cố, Free 100%, giảm all DV đơn lẻ salon)

BUG ĐÃ FIX:
  - Classifier match nhầm "ecom" trong "Shin-ECOM-bo" → dùng regex \becom\b word boundary.

MANUAL OVERRIDES (các quyết định thủ công user đã duyệt, cần giữ cho các tháng sau):
  - Campaign "Tôi yêu 30Shine ... Q2/Q3/..." → MEMBER (không phải T4 dù có "30Shine" trong tên)
  - Campaign #5041 (Combo 3,4,UND bill 500k+ tặng tinh dầu Arren date ngắn) → T4
  - Campaign "hàng cận date" → T4 (là CTKM đẩy SP cận date, không phải "ưu đãi khác")
  - Campaign "[CAMPAIGN KHÔNG GIA HẠN] Giảm 50k cho 3 lần tiếp theo [SC]" → VOUCHER (là thẻ giảm, đi cùng MKT)
  - Campaign "Shinecombo khách hàng BHXH" → MEMBER
  - Campaign "[Shine Member] Giảm 10% Shinecombo, 20% Cắt" → MEMBER
  - Campaign "1046/1047/1048 - 3 buổi Shinecombo N (HSD 6 tháng)" → MEMBER (thẻ dịch vụ)

TEXT CẦN GIỮ NGUYÊN (đã duyệt):
  - Summary button: "Xem X salon đã áp dụng campaign tháng N ▾"
  - KPI cards: 4 ô (CTKM đang áp dụng | Salon có phát sinh | Sắp hết hạn ≤10 ngày | Tổng lượt)
  - Header sub: "Cập nhật: DD/MM/YYYY · Dùng cho nhân sự Salon theo dõi"
    KHÔNG ghi ra "(đã loại NQ, MOYO)" hay "Chỉ tính Salon Công ty" — user muốn link chính thức sạch.

PUSH GITHUB:
  git remote origin = https://github.com/30shine-salon/ctkm-t4-30shine.git
  → push origin main (default, KHÔNG push lên quydo30shine trừ khi user yêu cầu rõ).
  Commit prefix: "bao-cao-ctkm-tN: <mô tả>"
"""
import argparse
import csv
import json
import re
import sys
from datetime import datetime
from html import escape
from pathlib import Path

import openpyxl

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')


# ============ CONFIG ============

CAT_LABELS = {
    'T4':      ('CTKM Trọng tâm Tháng {m}',         'or'),
    'SP':      ('Khuyến mãi Sản phẩm',              'nv'),
    'ECOM':    ('CTKM Ecom / Online',               'pur'),
    'MEMBER':  ('Ưu đãi Shine Member & Thẻ Dịch vụ','grn'),
    'VOUCHER': ('Voucher Code / MKT',               'nv'),
    'KHAC':    ('Ưu đãi khác / Salon riêng',        'dk'),
    'NOIBO':   ('Camp Nội bộ / Sự cố',              'dk'),
}
CAT_ORDER = ['T4', 'SP', 'ECOM', 'MEMBER', 'VOUCHER', 'KHAC', 'NOIBO']


def classify(name: str, month: int) -> str:
    """Phân loại campaign theo tên. Giữ thứ tự check này để tránh xung đột."""
    n = name.lower()

    # Ưu tiên cao nhất: các camp user đã gắn cứng về MEMBER
    if 'shine member' in n:
        return 'MEMBER'
    if 'shinecombo' in n and ('bhxh' in n or 'buổi' in n):
        return 'MEMBER'
    if 'tôi yêu 30shine' in n and 'moyo' not in n:
        return 'MEMBER'

    # T4 (hoặc tháng đang build)
    if f'ctkm tháng {month}' in n or f'ctkm t{month}' in n:
        return 'T4'
    if 'hàng cận date' in n:
        return 'T4'
    # Camp tặng tinh dầu date ngắn cho bill UND/Combo 3,4 (T4 trọng tâm)
    if 'date ngắn' in n and ('combo' in n or 'uốn' in n or 'nhuộm' in n):
        return 'T4'

    # Sản phẩm
    if 'laborie - combo' in n:
        return 'SP'
    if 'vu lan' in n:
        return 'SP'

    # ECOM (dùng word boundary để không match "shin-ecom-bo")
    if re.search(r'\becom\b', n) or n.startswith('ecom'):
        return 'ECOM'

    # MOYO - sẽ bị lọc ra sau nhưng vẫn gán để phân biệt
    if 'moyo' in n or 'chị đẹp' in n:
        return 'MOYO'

    # Thẻ dịch vụ N buổi
    if 'buổi' in n and ('hsd' in n or 'tháng' in n):
        return 'MEMBER'

    # Voucher / MKT
    if '[campaign không gia hạn]' in n and 'thẻ giảm' in n:
        return 'VOUCHER'
    if 'code' in n and ('dịch vụ' in n or 'giảm' in n):
        return 'VOUCHER'
    if 'voucher' in n or 'mkt' in n:
        return 'VOUCHER'

    # Nội bộ / sự cố
    if 'sự cố' in n or 'vh_' in n:
        return 'NOIBO'
    if ('giảm' in n and ('100%' in n or 'all dv' in n or 'all dịch vụ' in n)) and 'ctkm' not in n:
        return 'NOIBO'
    if 'free 100' in n or 'free dv' in n:
        return 'NOIBO'

    return 'KHAC'


def shorten(name: str):
    m = re.match(r'^\((\d+)\)\s*(.*)', name)
    if m:
        return m.group(1), m.group(2)
    return '', name


def load_allowed_salons(csv_path: Path) -> set:
    """Đọc danh sách salon công ty từ CSV phân cụm Salesup (cột 2)."""
    allow = set()
    with open(csv_path, 'r', encoding='utf-8') as f:
        r = csv.reader(f)
        next(r)  # skip header
        for row in r:
            if row and len(row) > 1 and row[1].strip():
                allow.add(row[1].strip())
    allow.add('Shop')  # Ecom shop cũng thuộc công ty
    return allow


def load_campaigns(xlsx_path: Path) -> dict:
    """Đọc data Excel. Header ở dòng 3, data bắt đầu dòng 4."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    camps = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 4 or not row[0]:
            continue
        name, start, end, salon, used = row
        if name not in camps:
            camps[name] = {'start': start, 'end': end, 'salons': [], 'total_used': 0}
        camps[name]['salons'].append({'name': salon, 'used': used or 0})
        camps[name]['total_used'] += (used or 0)
    return camps


def build_body(kept: dict, month: int) -> str:
    out = []
    for cat in CAT_ORDER:
        items = sorted(
            [(name, c) for name, c in kept.items() if c['category'] == cat],
            key=lambda x: (
                x[1].get('days_left', 9999) if x[1].get('days_left') is not None else 9999,
                -x[1]['total_used_filtered'],
            ),
        )
        if not items:
            continue
        label_tpl, color = CAT_LABELS[cat]
        label = label_tpl.format(m=month)
        total_used = sum(c['total_used_filtered'] for _, c in items)
        out.append(f'<!-- Category: {cat} -->')
        out.append(
            f'<div class="sec-hdg"><span class="cat-badge cat-{color}">{label}</span> '
            f'<span class="cat-cnt">{len(items)} CTKM · {total_used:,} lượt dùng</span></div>'
        )
        for name, c in items:
            code, short = shorten(name)
            ending_cls = ' ending-soon' if c.get('ending_soon') else ''
            end = c['end'].strftime('%d/%m/%Y')
            start = c['start'].strftime('%d/%m/%Y')
            salons = c['salons_filtered']
            n_sal = c['num_salons_filtered']
            total_used_f = c['total_used_filtered']
            salon_chips = ' '.join(
                f'<span class="salon-chip">{escape(s["name"])} <b>·{s["used"]}</b></span>'
                for s in salons[:60]
            )
            if n_sal > 60:
                salon_chips += f'<span class="salon-chip more">+{n_sal - 60} salon...</span>'
            dl = c.get('days_left')
            if dl is None:
                dl_badge = ''
            elif dl < 0:
                dl_badge = '<span class="dl-badge dl-exp">Hết</span>'
            elif dl == 0:
                dl_badge = '<span class="dl-badge dl-today">Hết hôm nay</span>'
            elif dl <= 7:
                dl_badge = f'<span class="dl-badge dl-soon">Còn {dl} ngày</span>'
            elif dl <= 30:
                dl_badge = f'<span class="dl-badge dl-mid">Còn {dl} ngày</span>'
            else:
                dl_badge = f'<span class="dl-badge dl-long">Còn {dl} ngày</span>'

            out.append('<div class="camp-card' + ending_cls + '">')
            out.append('  <div class="camp-head">')
            out.append(f'    <div class="camp-code">#{code}</div>')
            out.append(f'    <div class="camp-name">{escape(short)}</div>')
            out.append(f'    {dl_badge}')
            out.append('  </div>')
            out.append('  <div class="camp-meta">')
            out.append(f'    <span class="mi"><b>Thời gian:</b> {start} → {end}</span>')
            out.append(f'    <span class="mi"><b>Salon áp dụng:</b> {n_sal}</span>')
            out.append(f'    <span class="mi"><b>Đã dùng:</b> {total_used_f:,} lượt</span>')
            out.append('  </div>')
            if n_sal <= 3:
                out.append(f'  <div class="salon-list">{salon_chips}</div>')
            else:
                out.append(
                    f'  <details class="salon-det"><summary>Xem {n_sal} salon đã áp dụng campaign tháng {month} ▾</summary>'
                    f'<div class="salon-list">{salon_chips}</div></details>'
                )
            out.append('</div>')
    return '\n'.join(out)


CSS = r"""
:root{--navy:#1B3A6B;--navy-light:#2a5290;--orange:#FF7A29;--red:#E63946;--green:#2d8a4e;--green-lt:#e6f4ea;--yellow:#FFD23F;--cream:#FFF8EC;--bg:#f4f6f9;--card:#fff;--text:#1a1a2e;--muted:#6b7280;--border:#e5e7eb;--shadow:0 2px 12px rgba(27,58,107,.08);--r:14px;--purple:#7b1fa2}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Be Vietnam Pro',system-ui,sans-serif;background:var(--bg);color:var(--text);line-height:1.6;font-size:14px}
.hdr{background:linear-gradient(135deg,var(--navy),var(--navy-light));color:#fff;padding:16px 24px;position:sticky;top:0;z-index:100;border-bottom:4px solid var(--orange);display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px}
.hdr-left{display:flex;align-items:center;gap:14px}
.logo{width:46px;height:46px;background:var(--orange);border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:20px;font-weight:800;color:#fff}
.hdr h1{font-size:18px;font-weight:700}.hdr .sub{font-size:11px;opacity:.85;margin-top:2px}
.hdr-right{display:flex;gap:6px;align-items:center}
.badge{padding:4px 12px;border-radius:20px;font-size:11px;font-weight:700;letter-spacing:.4px;color:#fff}
.badge.or{background:var(--orange)}.badge.gr{background:var(--green)}
.wrap{max-width:1200px;margin:0 auto;padding:20px 16px 60px}
.kpi-g{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;margin-bottom:20px}
.kpi{background:#fff;border-radius:12px;padding:16px;text-align:center;border:1px solid var(--border);box-shadow:var(--shadow);border-top:3px solid var(--navy)}
.kpi.or{border-top-color:var(--orange)}.kpi.rd{border-top-color:var(--red)}
.kpi.gr{border-top-color:var(--green)}.kpi.pp{border-top-color:var(--purple)}
.kpi .v{font-size:28px;font-weight:800;color:var(--navy);line-height:1}
.kpi .l{font-size:11px;color:var(--muted);margin-top:6px;font-weight:500}
.intro{background:linear-gradient(135deg,#fff8ec,#fff2d9);border-left:4px solid var(--orange);border-radius:12px;padding:16px 20px;margin-bottom:22px;font-size:13px;line-height:1.7}
.intro b{color:var(--navy)}.intro ul{margin:6px 0 0 20px}
.sec-hdg{display:flex;align-items:center;justify-content:space-between;margin:28px 0 12px;padding:0;gap:10px;flex-wrap:wrap}
.cat-badge{display:inline-flex;align-items:center;font-size:14px;font-weight:800;color:#fff;padding:8px 16px;border-radius:10px;letter-spacing:.2px}
.cat-or{background:linear-gradient(135deg,#e65100,var(--orange))}
.cat-nv{background:linear-gradient(135deg,var(--navy),var(--navy-light))}
.cat-pur{background:linear-gradient(135deg,#4a148c,var(--purple))}
.cat-grn{background:linear-gradient(135deg,#1a5632,var(--green))}
.cat-dk{background:linear-gradient(135deg,#37474f,#607d8b)}
.cat-cnt{font-size:12px;color:var(--muted);font-weight:600}
.camp-card{background:#fff;border-radius:12px;padding:14px 16px;margin-bottom:10px;border:1px solid var(--border);box-shadow:0 1px 4px rgba(0,0,0,.04);transition:.15s}
.camp-card:hover{box-shadow:0 4px 14px rgba(27,58,107,.1);border-color:#d0d7e2}
.camp-card.ending-soon{border-left:4px solid var(--red);background:linear-gradient(90deg,#fff8f8,#fff)}
.camp-head{display:flex;align-items:flex-start;gap:10px;margin-bottom:6px}
.camp-code{background:var(--navy);color:#fff;font-size:11px;font-weight:700;padding:3px 9px;border-radius:6px;font-family:ui-monospace,monospace;flex-shrink:0;letter-spacing:.3px}
.camp-name{flex:1;font-size:13.5px;font-weight:600;color:var(--text);line-height:1.45}
.dl-badge{font-size:11px;font-weight:700;padding:3px 10px;border-radius:20px;flex-shrink:0;white-space:nowrap}
.dl-exp{background:#e5e7eb;color:#6b7280}
.dl-today{background:var(--red);color:#fff;animation:pulse 1.5s infinite}
.dl-soon{background:#fee;color:var(--red);border:1px solid #fcc}
.dl-mid{background:#fff3cd;color:#856404}
.dl-long{background:var(--green-lt);color:var(--green)}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.7}}
.camp-meta{display:flex;flex-wrap:wrap;gap:4px 18px;font-size:12px;color:var(--muted);padding-left:2px;margin-bottom:6px}
.camp-meta .mi b{color:var(--text);font-weight:600}
.salon-det{margin-top:6px}
.salon-det summary{cursor:pointer;font-size:11.5px;color:var(--navy);font-weight:600;padding:3px 0;user-select:none;list-style:none}
.salon-det summary::-webkit-details-marker{display:none}
.salon-det summary:hover{color:var(--orange)}
.salon-det[open] summary{margin-bottom:6px}
.salon-list{display:flex;flex-wrap:wrap;gap:5px;margin-top:4px;padding:6px 0}
.salon-chip{background:#f0f4ff;border:1px solid #d5deef;color:var(--navy);font-size:11px;padding:2px 8px;border-radius:12px;font-weight:500;line-height:1.6}
.salon-chip b{color:var(--orange);font-weight:700;margin-left:2px}
.salon-chip.more{background:#f3e8ff;border-color:#e6d4ff;color:var(--purple)}
@media print{.hdr{position:static}.camp-card{break-inside:avoid}.salon-det{display:none}body{font-size:11px}}
@media(max-width:640px){.hdr h1{font-size:15px}.camp-head{flex-wrap:wrap}.kpi .v{font-size:22px}}
"""


def render_html(body: str, stats: dict, today: datetime, month: int) -> str:
    today_str = today.strftime('%d/%m/%Y')
    return f"""<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Báo cáo Rà soát CTKM đang hoạt động — Tháng {month}/{today.year}</title>
<link href="https://fonts.googleapis.com/css2?family=Be+Vietnam+Pro:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>{CSS}</style>
</head>
<body>

<div class="hdr">
  <div class="hdr-left">
    <div class="logo">30</div>
    <div>
      <h1>Rà soát CTKM đang hoạt động — Tháng {month}/{today.year}</h1>
      <div class="sub">Cập nhật: {today_str} · Dùng cho nhân sự Salon theo dõi</div>
    </div>
  </div>
  <div class="hdr-right">
    <span class="badge or">{stats['total_kept']} CTKM</span>
    <span class="badge gr">{stats['total_salons']} Salon</span>
  </div>
</div>

<div class="wrap">

  <div class="kpi-g">
    <div class="kpi"><div class="v">{stats['total_kept']}</div><div class="l">CTKM đang áp dụng</div></div>
    <div class="kpi or"><div class="v">{stats['total_salons']}</div><div class="l">Salon có phát sinh áp dụng</div></div>
    <div class="kpi rd"><div class="v">{stats['total_ending_soon']}</div><div class="l">Sắp hết hạn (≤ 10 ngày)</div></div>
    <div class="kpi gr"><div class="v">{stats['total_used']:,}</div><div class="l">Tổng lượt khách đã áp dụng</div></div>
  </div>

  <div class="intro">
    <b>Hướng dẫn đọc báo cáo:</b>
    <ul>
      <li>CTKM được nhóm theo <b>mức độ ưu tiên truyền thông</b> — Nhóm 1 (CTKM Trọng tâm T{month}) là nhóm salon phải thuộc kịch bản tư vấn.</li>
      <li>Thẻ đỏ <span class="dl-badge dl-soon" style="margin:0 3px">Còn X ngày</span> = sắp hết hạn (≤ 7 ngày), cần tận dụng tối đa trước khi kết thúc.</li>
      <li>Click <b>"Xem X salon đã áp dụng campaign tháng {month} ▾"</b> để mở danh sách chi tiết salon + số lượt đã dùng.</li>
    </ul>
  </div>

  {body}

  <div style="text-align:center;color:var(--muted);font-size:11px;margin-top:30px;padding-top:20px;border-top:1px solid var(--border)">
    © 30Shine · Nguồn: <i>data campaign đang hoạt động T{month}</i> + <i>Chia salon cụm Salesup T{month}</i> · Generated {today_str}
  </div>

</div>

</body>
</html>
"""


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--month', type=int, default=datetime.now().month)
    p.add_argument('--today', type=str, default=datetime.now().strftime('%Y-%m-%d'),
                   help='Ngày chốt báo cáo YYYY-MM-DD (camp có end >= today mới được giữ)')
    p.add_argument('--xlsx', type=str, default=None, help='File Excel (mặc định: data campaign dang hoat dong tN.xlsx)')
    p.add_argument('--csv',  type=str, default=None, help='File CSV phân cụm salon (mặc định: Chia salon cụm Salesup - phân cụm mới tháng N.csv)')
    p.add_argument('--out',  type=str, default=None, help='File HTML output (mặc định: bao-cao-ctkm-dang-hoat-dong-tN.html)')
    args = p.parse_args()

    month = args.month
    today = datetime.strptime(args.today, '%Y-%m-%d')
    xlsx = Path(args.xlsx or f'data campaign dang hoat dong t{month}.xlsx')
    csv_path = Path(args.csv or f'Chia salon cụm Salesup - phân cụm mới tháng {month}.csv')
    out = Path(args.out or f'bao-cao-ctkm-dang-hoat-dong-t{month}.html')

    if not xlsx.exists():
        sys.exit(f'Không tìm thấy file Excel: {xlsx}')
    if not csv_path.exists():
        sys.exit(f'Không tìm thấy file CSV phân cụm: {csv_path}')

    allow = load_allowed_salons(csv_path)
    camps = load_campaigns(xlsx)

    for name, c in camps.items():
        c['category'] = classify(name, month)
        c['active'] = c['end'] and c['end'] >= today
        c['days_left'] = (c['end'] - today).days if c['end'] else None
        c['ending_soon'] = c['days_left'] is not None and 0 <= c['days_left'] <= 10
        filtered = [s for s in c['salons'] if s['name'] in allow]
        c['salons_filtered'] = sorted(filtered, key=lambda x: -x['used'])
        c['num_salons_filtered'] = len(filtered)
        c['total_used_filtered'] = sum(s['used'] for s in filtered)

    # Bỏ MOYO, camp hết hạn, camp không còn salon công ty
    kept = {
        name: c for name, c in camps.items()
        if c['active'] and c['category'] != 'MOYO' and c['num_salons_filtered'] > 0
        and not name.startswith('(5074)')  # Tôi yêu 30Shine - Moyo
    }

    stats = {
        'total_kept': len(kept),
        'total_ending_soon': sum(1 for c in kept.values() if c.get('ending_soon')),
        'total_used': sum(c['total_used_filtered'] for c in kept.values()),
        'total_salons': len({s['name'] for c in kept.values() for s in c['salons_filtered']}),
    }

    body = build_body(kept, month)
    html = render_html(body, stats, today, month)
    out.write_text(html, encoding='utf-8')

    from collections import Counter
    per_cat = Counter(c['category'] for c in kept.values())
    print(f'[OK] Generated {out}')
    print(f'  Kept: {stats["total_kept"]} / {len(camps)} CTKM')
    print(f'  Per category: {dict(per_cat)}')
    print(f'  {stats["total_used"]:,} luot | {stats["total_salons"]} salon | {stats["total_ending_soon"]} sap het han')


if __name__ == '__main__':
    main()
