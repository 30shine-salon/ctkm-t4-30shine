import csv
import io
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

# --- Read CSV ---
rows = []
with open(r"d:\Downloads\CTKM t4\CTKM tháng 3 - Sheet1.csv", encoding="utf-8") as f:
    reader = csv.reader(f)
    header = next(reader)
    for r in reader:
        if len(r) >= 6 and r[0].strip():
            campaign = r[0].strip()
            start = r[1].strip()
            end = r[2].strip()
            salon = r[3].strip()
            try:
                used = int(r[4].strip())
            except:
                continue
            category = r[5].strip()
            rows.append({
                'campaign': campaign,
                'start': start,
                'end': end,
                'salon': salon,
                'used': used,
                'category': category
            })

# --- Styles ---
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBHEADER_FONT = Font(bold=True, size=11)
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFF2CC")
BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
TITLE_FONT = Font(bold=True, size=14, color="2F5496")
SECTION_FONT = Font(bold=True, size=12, color="2F5496")

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER

def style_data_rows(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='center')

def auto_width(ws, cols, max_width=45):
    for c in range(1, cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 3, max_width)

def write_table(ws, start_row, headers, data_rows):
    """Write a table with headers and data, return next empty row."""
    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i, value=h)
    style_header(ws, start_row, len(headers))
    for ri, dr in enumerate(data_rows, start_row + 1):
        for ci, val in enumerate(dr, 1):
            ws.cell(row=ri, column=ci, value=val)
    style_data_rows(ws, start_row + 1, start_row + len(data_rows), len(headers))
    return start_row + len(data_rows) + 2

# --- Read Salon Master List ---
salon_master = []
with open(r"d:\Downloads\CTKM t4\Chia salon cụm Salesup - phân cụm mới tháng 4.csv", encoding="utf-8") as f:
    reader = csv.reader(f)
    header_salon = next(reader)
    for r in reader:
        if len(r) >= 5 and r[1].strip():
            salon_master.append({
                'id': r[0].strip(),
                'name': r[1].strip(),
                'supervisor': r[2].strip(),
                'cluster': r[3].strip(),
                'type': r[4].strip(),
            })

# ===================== CẤU HÌNH 3 NHÓM CTKM =====================
# Nhóm A: SP Mới về / Mẫu mã mới   → Chưa có lịch sử, đo trial & awareness
# Nhóm B: SP Thường kích nhu cầu    → Đo % KH mới, CAC, repeat
# Nhóm C: Tồn kho date dài          → Đo tốc độ xả, % đã xả, burn rate
# (Clearstock có rule riêng, KHÔNG đánh giá ở đây)

GROUP_RULES = {
    'A': {
        'name': 'SP Mới về',
        'margin_rate_min': 30,
        'giam_gia_max': 20,
        'do_phu_salon_min': 50,
        'roi_min': 80,
    },
    'B': {
        'name': 'SP Thường kích nhu cầu',
        'margin_rate_min': 40,
        'giam_gia_max': 30,
        'pct_kh_moi_min': 70,
        'cac_multiplier': 2,     # CAC tối đa = 2× margin/lượt
    },
    'C': {
        'name': 'Tồn kho date dài',
        'margin_rate_min': 15,
        'giam_gia_max': 50,
        'do_phu_salon_min': 80,
    },
}

# --- CẤU HÌNH SẢN PHẨM ---
# keyword: tìm trong tên campaign (lowercase), khớp campaign nào chứa keyword
# gia_goc, cogs, gia_km: đơn vị VND
# pct_kh_moi: % KH mới ước tính (chỉ Nhóm B)
# ton_kho_ban_dau, ton_kho_con_lai, ngay_het_date: chỉ Nhóm C
PRODUCT_CONFIG = [
    {
        'keyword': 'laborie',
        'name': 'Laborie',
        'group': 'B',
        'gia_goc': 400_000,
        'cogs': 130_000,
        'gia_km': 280_000,
        'pct_kh_moi': 65,
        'ton_kho_ban_dau': None,
        'ton_kho_con_lai': None,
        'ngay_het_date': None,
    },
    {
        'keyword': 'máy sấy',
        'name': 'Máy sấy tóc',
        'group': 'A',
        'gia_goc': 700_000,
        'cogs': 350_000,
        'gia_km': 605_000,
        'pct_kh_moi': None,
        'ton_kho_ban_dau': None,
        'ton_kho_con_lai': None,
        'ngay_het_date': None,
    },
    {
        'keyword': 'massage',
        'name': 'Máy Massage',
        'group': 'C',
        'gia_goc': 1_200_000,
        'cogs': 500_000,
        'gia_km': 600_000,
        'pct_kh_moi': None,
        'ton_kho_ban_dau': 200,
        'ton_kho_con_lai': 140,
        'ngay_het_date': '2026-09-30',
    },
    {
        'keyword': 'tôi yêu 30shine',
        'name': 'Tôi yêu 30Shine (SP chung)',
        'group': 'B',
        'gia_goc': 333_000,
        'cogs': 100_000,
        'gia_km': 233_000,
        'pct_kh_moi': 70,
        'ton_kho_ban_dau': None,
        'ton_kho_con_lai': None,
        'ngay_het_date': None,
    },
    {
        'keyword': 'glanzen',
        'name': 'Sáp Glanzen Prime Sandalwood',
        'group': 'B',
        'gia_goc': 279_000,
        'cogs': 75_000,
        'gia_km': 223_200,       # giảm 20%
        'pct_kh_moi': 70,
        'ton_kho_ban_dau': None,
        'ton_kho_con_lai': None,
        'ngay_het_date': None,
    },
    {
        'keyword': 'dabo',
        'name': 'Serum Dabo 7in1 (tặng tẩy da chết Apteka)',
        'group': 'B',
        'gia_goc': 599_000,
        'cogs': 217_000,        # 180k SP + 37k quà tặng Apteka
        'gia_km': 599_000,      # không giảm giá, tặng kèm
        'pct_kh_moi': 70,
        'ton_kho_ban_dau': None,
        'ton_kho_con_lai': None,
        'ngay_het_date': None,
    },
    # Thêm SP khác tại đây...
]

# ===================== CALCULATIONS =====================

# --- Overall ---
total_used = sum(r['used'] for r in rows)
unique_campaigns = list(set(r['campaign'] for r in rows))
unique_salons = list(set(r['salon'] for r in rows))

# By category
cat_stats = defaultdict(lambda: {'records': 0, 'used': 0, 'salons': set(), 'campaigns': set()})
for r in rows:
    cat_stats[r['category']]['records'] += 1
    cat_stats[r['category']]['used'] += r['used']
    cat_stats[r['category']]['salons'].add(r['salon'])
    cat_stats[r['category']]['campaigns'].add(r['campaign'])

# Top campaigns
camp_stats = defaultdict(lambda: {'used': 0, 'salons': set(), 'category': ''})
for r in rows:
    camp_stats[r['campaign']]['used'] += r['used']
    camp_stats[r['campaign']]['salons'].add(r['salon'])
    camp_stats[r['campaign']]['category'] = r['category']
    camp_stats[r['campaign']]['start'] = r['start']
    camp_stats[r['campaign']]['end'] = r['end']

top_campaigns = sorted(camp_stats.items(), key=lambda x: -x[1]['used'])[:15]

# Top salons
salon_stats = defaultdict(lambda: {'used': 0, 'campaigns': set(), 'categories': set()})
for r in rows:
    salon_stats[r['salon']]['used'] += r['used']
    salon_stats[r['salon']]['campaigns'].add(r['campaign'])
    salon_stats[r['salon']]['categories'].add(r['category'])
top_salons = sorted(salon_stats.items(), key=lambda x: -x[1]['used'])[:15]

# ===================== CLEARSTOCK deep dive =====================
cs_rows = [r for r in rows if r['category'] == 'Clearstock']
cs_total = sum(r['used'] for r in cs_rows)

# Clearstock by campaign
cs_camp = defaultdict(lambda: {'used': 0, 'salons': set(), 'start': '', 'end': ''})
for r in cs_rows:
    cs_camp[r['campaign']]['used'] += r['used']
    cs_camp[r['campaign']]['salons'].add(r['salon'])
    cs_camp[r['campaign']]['start'] = r['start']
    cs_camp[r['campaign']]['end'] = r['end']
cs_camp_sorted = sorted(cs_camp.items(), key=lambda x: -x[1]['used'])

# Clearstock by salon
cs_salon = defaultdict(lambda: {'used': 0, 'campaigns': set()})
for r in cs_rows:
    cs_salon[r['salon']]['used'] += r['used']
    cs_salon[r['salon']]['campaigns'].add(r['campaign'])
cs_salon_sorted = sorted(cs_salon.items(), key=lambda x: -x[1]['used'])

# Clearstock: salon x campaign matrix
cs_matrix_salons = [s[0] for s in cs_salon_sorted]
cs_matrix_camps = [c[0] for c in cs_camp_sorted]

# ===================== CTKM Tháng 3 deep dive =====================
ct_rows = [r for r in rows if r['category'] == 'CTKM Tháng 3']
ct_total = sum(r['used'] for r in ct_rows)

ct_camp = defaultdict(lambda: {'used': 0, 'salons': set(), 'start': '', 'end': ''})
for r in ct_rows:
    ct_camp[r['campaign']]['used'] += r['used']
    ct_camp[r['campaign']]['salons'].add(r['salon'])
    ct_camp[r['campaign']]['start'] = r['start']
    ct_camp[r['campaign']]['end'] = r['end']
ct_camp_sorted = sorted(ct_camp.items(), key=lambda x: -x[1]['used'])

ct_salon = defaultdict(lambda: {'used': 0, 'campaigns': set()})
for r in ct_rows:
    ct_salon[r['salon']]['used'] += r['used']
    ct_salon[r['salon']]['campaigns'].add(r['campaign'])
ct_salon_sorted = sorted(ct_salon.items(), key=lambda x: -x[1]['used'])

# ===================== Salon analysis for both categories =====================
# For each salon: clearstock used, ctkm used, total, campaigns list
all_salons_both = set()
for r in rows:
    if r['category'] in ('Clearstock', 'CTKM Tháng 3'):
        all_salons_both.add(r['salon'])

salon_both = {}
for s in all_salons_both:
    salon_both[s] = {
        'cs_used': 0, 'ct_used': 0, 'cs_camps': set(), 'ct_camps': set(),
        'total': 0, 'cs_records': 0, 'ct_records': 0
    }
for r in rows:
    if r['salon'] in all_salons_both:
        if r['category'] == 'Clearstock':
            salon_both[r['salon']]['cs_used'] += r['used']
            salon_both[r['salon']]['cs_camps'].add(r['campaign'])
            salon_both[r['salon']]['cs_records'] += 1
        elif r['category'] == 'CTKM Tháng 3':
            salon_both[r['salon']]['ct_used'] += r['used']
            salon_both[r['salon']]['ct_camps'].add(r['campaign'])
            salon_both[r['salon']]['ct_records'] += 1
        salon_both[r['salon']]['total'] = salon_both[r['salon']]['cs_used'] + salon_both[r['salon']]['ct_used']

salon_both_sorted = sorted(salon_both.items(), key=lambda x: -x[1]['total'])

# ===================== ĐÁNH GIÁ CTKM THEO 3 NHÓM =====================
non_cs_rows = [r for r in rows if r['category'] != 'Clearstock']
salon_master_count = len(salon_master)

product_eval = []
for pc in PRODUCT_CONFIG:
    kw = pc['keyword'].lower()
    # Tìm campaign khớp keyword (bỏ Clearstock)
    match_camp = defaultdict(lambda: {'used': 0, 'salons': set(), 'start': '', 'end': ''})
    for r in non_cs_rows:
        if kw in r['campaign'].lower():
            match_camp[r['campaign']]['used'] += r['used']
            match_camp[r['campaign']]['salons'].add(r['salon'])
            match_camp[r['campaign']]['start'] = r['start']
            match_camp[r['campaign']]['end'] = r['end']

    if not match_camp:
        continue

    total_sd = sum(c['used'] for c in match_camp.values())
    all_sl = set()
    for c in match_camp.values():
        all_sl.update(c['salons'])
    num_salons = len(all_sl)

    # Thời gian KM: trung bình có trọng số theo lượt SD của từng campaign
    # (tránh bị lệch bởi 1 campaign date rất dài)
    weighted_days = 0
    weight_total = 0
    for c in match_camp.values():
        try:
            s = datetime.strptime(c['start'], '%Y-%m-%d')
            e = datetime.strptime(c['end'], '%Y-%m-%d')
            d = max((e - s).days, 1)
            if d > 365:  # bỏ campaign date bất thường (> 1 năm)
                continue
            weighted_days += d * c['used']
            weight_total += c['used']
        except:
            pass
    duration = round(weighted_days / weight_total) if weight_total > 0 else 30

    # --- Tính toán core ---
    gia_goc = pc['gia_goc']
    cogs = pc['cogs']
    gia_km = pc['gia_km']
    margin = gia_km - cogs                          # Lợi nhuận/lượt
    discount = gia_goc - gia_km                     # Mức giảm/lượt
    discount_pct = round(discount / gia_goc * 100, 1) if gia_goc else 0
    margin_rate = round(margin / gia_km * 100, 1) if gia_km else 0
    total_margin = margin * total_sd
    total_cost_km = discount * total_sd
    roi = round(total_margin / total_cost_km * 100, 1) if total_cost_km > 0 else 0
    do_phu = round(num_salons / salon_master_count * 100, 1) if salon_master_count else 0

    group = pc['group']
    rules = GROUP_RULES[group]

    ev = {
        'name': pc['name'], 'group': group, 'group_name': rules['name'],
        'gia_goc': gia_goc, 'cogs': cogs, 'gia_km': gia_km,
        'margin': margin, 'discount': discount, 'discount_pct': discount_pct,
        'margin_rate': margin_rate, 'total_sd': total_sd, 'num_salons': num_salons,
        'duration': duration, 'total_margin': total_margin,
        'total_cost_km': total_cost_km, 'roi': roi, 'do_phu': do_phu,
        'campaigns': list(match_camp.keys()),
        'checks': [], 'recs': [],
    }

    checks = []

    # --- Check chung: Margin Rate ---
    checks.append({
        'rule': f"Margin Rate >= {rules['margin_rate_min']}%",
        'actual': f"{margin_rate}%",
        'target': f">= {rules['margin_rate_min']}%",
        'ok': margin_rate >= rules['margin_rate_min'],
    })
    # --- Check chung: Giảm giá ---
    checks.append({
        'rule': f"Giảm giá <= {rules['giam_gia_max']}%",
        'actual': f"{discount_pct}%",
        'target': f"<= {rules['giam_gia_max']}%",
        'ok': discount_pct <= rules['giam_gia_max'],
    })

    # --- NHÓM A ---
    if group == 'A':
        checks.append({
            'rule': f"ROI >= {rules['roi_min']}%",
            'actual': f"{roi}%",
            'target': f">= {rules['roi_min']}%",
            'ok': roi >= rules['roi_min'],
        })
        checks.append({
            'rule': f"Độ phủ salon >= {rules['do_phu_salon_min']}%",
            'actual': f"{do_phu}%",
            'target': f">= {rules['do_phu_salon_min']}%",
            'ok': do_phu >= rules['do_phu_salon_min'],
        })
        # Đề xuất
        if do_phu < rules['do_phu_salon_min']:
            ev['recs'].append(f"Mở rộng thêm salon (hiện {do_phu}%, cần >= {rules['do_phu_salon_min']}%)")

    # --- NHÓM B ---
    elif group == 'B':
        pct_kh_moi = pc.get('pct_kh_moi') or 0
        so_kh_moi = round(total_sd * pct_kh_moi / 100) if pct_kh_moi else 0
        cac = round(total_cost_km / so_kh_moi) if so_kh_moi > 0 else 0
        cac_limit = margin * rules.get('cac_multiplier', 2)
        payback = round(cac / margin, 1) if margin > 0 else 0
        ev['pct_kh_moi'] = pct_kh_moi
        ev['so_kh_moi'] = so_kh_moi
        ev['cac'] = cac
        ev['cac_limit'] = cac_limit
        ev['payback'] = payback

        checks.append({
            'rule': f"% KH mới >= {rules['pct_kh_moi_min']}%",
            'actual': f"{pct_kh_moi}%",
            'target': f">= {rules['pct_kh_moi_min']}%",
            'ok': pct_kh_moi >= rules['pct_kh_moi_min'],
        })
        checks.append({
            'rule': f"CAC <= {rules['cac_multiplier']}x Margin ({cac_limit:,.0f}đ)",
            'actual': f"{cac:,.0f}đ",
            'target': f"<= {cac_limit:,.0f}đ",
            'ok': cac <= cac_limit,
        })
        # Đề xuất
        if pct_kh_moi < rules['pct_kh_moi_min']:
            ev['recs'].append(f"Nhắm target KH mới nhiều hơn (hiện {pct_kh_moi}%, cần >= {rules['pct_kh_moi_min']}%)")
        if cac > cac_limit:
            ev['recs'].append("Giảm mức KM hoặc tăng giá KM để giảm CAC")
        if margin_rate < rules['margin_rate_min']:
            ev['recs'].append(f"Tăng giá KM hoặc giảm COGS để đạt Margin Rate >= {rules['margin_rate_min']}%")

    # --- NHÓM C ---
    elif group == 'C':
        ton_kho = pc.get('ton_kho_ban_dau') or 0
        ton_con = pc.get('ton_kho_con_lai') or 0
        ngay_het = pc.get('ngay_het_date')

        pct_xa = round((ton_kho - ton_con) / ton_kho * 100, 1) if ton_kho > 0 else 0
        vong_quay = round(gia_km * total_sd / (cogs * ton_kho), 2) if ton_kho > 0 and cogs > 0 else 0
        burn_rate_week = round(total_sd / max(duration / 7, 1), 1)

        ev['pct_xa'] = pct_xa
        ev['vong_quay'] = vong_quay
        ev['burn_rate_week'] = burn_rate_week
        ev['ton_kho'] = ton_kho
        ev['ton_con'] = ton_con

        checks.append({
            'rule': f"Độ phủ salon >= {rules['do_phu_salon_min']}%",
            'actual': f"{do_phu}%",
            'target': f">= {rules['do_phu_salon_min']}%",
            'ok': do_phu >= rules['do_phu_salon_min'],
        })

        if ngay_het:
            try:
                date_exp = datetime.strptime(ngay_het, '%Y-%m-%d')
                days_left = (date_exp - datetime.now()).days
                ev['days_left'] = days_left
                if burn_rate_week > 0 and ton_con > 0:
                    weeks_need = round(ton_con / burn_rate_week, 1)
                    weeks_avail = round(days_left / 7, 1)
                    ev['weeks_need'] = weeks_need
                    ev['weeks_avail'] = weeks_avail
            except:
                pass

        # Đề xuất
        if do_phu < rules['do_phu_salon_min']:
            ev['recs'].append(f"Mở rộng salon (hiện {do_phu}%, cần >= {rules['do_phu_salon_min']}%)")

    # --- Kết quả chung ---
    ev['checks'] = checks
    ev['pass_count'] = sum(1 for c in checks if c['ok'])
    ev['total_checks'] = len(checks)
    ev['pass_pct'] = round(ev['pass_count'] / ev['total_checks'] * 100) if ev['total_checks'] > 0 else 0

    if ev['pass_pct'] >= 80:
        ev['verdict'] = 'ĐẠT'
    elif ev['pass_pct'] >= 50:
        ev['verdict'] = 'CẦN CẢI THIỆN'
    else:
        ev['verdict'] = 'KHÔNG ĐẠT'

    if not ev['recs']:
        ev['recs'].append("Đạt các chỉ tiêu, tiếp tục theo dõi")

    product_eval.append(ev)

# ===================== CREATE WORKBOOK =====================
wb = Workbook()

# ============ SHEET 1: TỔNG QUAN ============
ws1 = wb.active
ws1.title = "Tổng quan"

ws1.cell(row=1, column=1, value="BÁO CÁO TỔNG QUAN CTKM THÁNG 3 - 30SHINE").font = TITLE_FONT
ws1.merge_cells('A1:F1')

# KPI summary
ws1.cell(row=3, column=1, value="CHỈ SỐ TỔNG QUAN").font = SECTION_FONT
kpi_headers = ["Chỉ số", "Giá trị"]
kpi_data = [
    ["Tổng số bản ghi", len(rows)],
    ["Tổng chiến dịch", len(unique_campaigns)],
    ["Tổng salon", len(unique_salons)],
    ["Tổng lượt sử dụng", total_used],
    ["Trung bình lượt SD / bản ghi", round(total_used / len(rows), 2)],
    ["Trung bình lượt SD / chiến dịch", round(total_used / len(unique_campaigns), 2)],
    ["Trung bình lượt SD / salon", round(total_used / len(unique_salons), 2)],
]
r = write_table(ws1, 4, kpi_headers, kpi_data)

# Category breakdown
ws1.cell(row=r, column=1, value="PHÂN BỔ THEO DANH MỤC").font = SECTION_FONT
cat_headers = ["Danh mục", "Số bản ghi", "Tổng SD", "% Tổng SD", "Số chiến dịch", "Số salon", "TB SD/bản ghi"]
cat_data = []
for cat in sorted(cat_stats.keys(), key=lambda x: -cat_stats[x]['used']):
    s = cat_stats[cat]
    cat_data.append([
        cat, s['records'], s['used'],
        round(s['used'] / total_used * 100, 1),
        len(s['campaigns']), len(s['salons']),
        round(s['used'] / s['records'], 2)
    ])
r = write_table(ws1, r + 1, cat_headers, cat_data)

# Top campaigns
ws1.cell(row=r, column=1, value="TOP 15 CHIẾN DỊCH THEO LƯỢT SỬ DỤNG").font = SECTION_FONT
tc_headers = ["Chiến dịch", "Danh mục", "Bắt đầu", "Kết thúc", "Số salon", "Tổng SD", "% Tổng"]
tc_data = []
for name, info in top_campaigns:
    tc_data.append([
        name, info['category'], info['start'], info['end'],
        len(info['salons']), info['used'],
        round(info['used'] / total_used * 100, 1)
    ])
r = write_table(ws1, r + 1, tc_headers, tc_data)

# Top salons
ws1.cell(row=r, column=1, value="TOP 15 SALON THEO LƯỢT SỬ DỤNG").font = SECTION_FONT
ts_headers = ["Salon", "Tổng SD", "% Tổng", "Số chiến dịch", "Các danh mục"]
ts_data = []
for name, info in top_salons:
    ts_data.append([
        name, info['used'],
        round(info['used'] / total_used * 100, 1),
        len(info['campaigns']),
        ", ".join(sorted(info['categories']))
    ])
r = write_table(ws1, r + 1, ts_headers, ts_data)

auto_width(ws1, 7)

# ============ SHEET 2: CLEARSTOCK CHI TIẾT ============
ws2 = wb.create_sheet("Clearstock - Chi tiết")

ws2.cell(row=1, column=1, value="PHÂN TÍCH SÂU: CLEARSTOCK").font = TITLE_FONT
ws2.merge_cells('A1:F1')

# KPI
ws2.cell(row=3, column=1, value="TỔNG QUAN CLEARSTOCK").font = SECTION_FONT
cs_kpi = [
    ["Tổng bản ghi", len(cs_rows)],
    ["Tổng lượt sử dụng", cs_total],
    ["% so với tổng CTKM", f"{round(cs_total / total_used * 100, 1)}%"],
    ["Số chiến dịch", len(cs_camp)],
    ["Số salon tham gia", len(cs_salon)],
    ["TB lượt SD / bản ghi", round(cs_total / len(cs_rows), 2)],
    ["TB lượt SD / chiến dịch", round(cs_total / len(cs_camp), 2)],
    ["TB lượt SD / salon", round(cs_total / len(cs_salon), 2)],
]
r = write_table(ws2, 4, ["Chỉ số", "Giá trị"], cs_kpi)

# By campaign
ws2.cell(row=r, column=1, value="CHI TIẾT THEO CHIẾN DỊCH CLEARSTOCK").font = SECTION_FONT
cs_c_headers = ["Chiến dịch", "Bắt đầu", "Kết thúc", "Số salon", "Tổng SD", "% Clearstock", "TB SD/salon"]
cs_c_data = []
for name, info in cs_camp_sorted:
    ns = len(info['salons'])
    cs_c_data.append([
        name, info['start'], info['end'], ns, info['used'],
        round(info['used'] / cs_total * 100, 1),
        round(info['used'] / ns, 2) if ns else 0
    ])
r = write_table(ws2, r + 1, cs_c_headers, cs_c_data)

# By salon
ws2.cell(row=r, column=1, value="CHI TIẾT THEO SALON - CLEARSTOCK").font = SECTION_FONT
cs_s_headers = ["Salon", "Tổng SD", "% Clearstock", "Số chiến dịch", "TB SD/chiến dịch", "Danh sách chiến dịch"]
cs_s_data = []
for name, info in cs_salon_sorted:
    nc = len(info['campaigns'])
    cs_s_data.append([
        name, info['used'],
        round(info['used'] / cs_total * 100, 1),
        nc,
        round(info['used'] / nc, 2) if nc else 0,
        "; ".join(sorted(info['campaigns']))[:200]
    ])
r = write_table(ws2, r + 1, cs_s_headers, cs_s_data)

# Raw data
ws2.cell(row=r, column=1, value="DỮ LIỆU CHI TIẾT CLEARSTOCK").font = SECTION_FONT
raw_headers = ["Chiến dịch", "Bắt đầu", "Kết thúc", "Salon", "Đã sử dụng"]
raw_data = [[row['campaign'], row['start'], row['end'], row['salon'], row['used']] for row in cs_rows]
raw_data.sort(key=lambda x: -x[4])
r = write_table(ws2, r + 1, raw_headers, raw_data)

auto_width(ws2, 7)

# ============ SHEET 3: CTKM THÁNG 3 CHI TIẾT ============
ws3 = wb.create_sheet("CTKM Tháng 3 - Chi tiết")

ws3.cell(row=1, column=1, value="PHÂN TÍCH SÂU: CTKM THÁNG 3").font = TITLE_FONT
ws3.merge_cells('A1:F1')

ws3.cell(row=3, column=1, value="TỔNG QUAN CTKM THÁNG 3").font = SECTION_FONT
ct_kpi = [
    ["Tổng bản ghi", len(ct_rows)],
    ["Tổng lượt sử dụng", ct_total],
    ["% so với tổng CTKM", f"{round(ct_total / total_used * 100, 1)}%"],
    ["Số chiến dịch", len(ct_camp)],
    ["Số salon tham gia", len(ct_salon)],
    ["TB lượt SD / bản ghi", round(ct_total / len(ct_rows), 2) if ct_rows else 0],
    ["TB lượt SD / chiến dịch", round(ct_total / len(ct_camp), 2) if ct_camp else 0],
    ["TB lượt SD / salon", round(ct_total / len(ct_salon), 2) if ct_salon else 0],
]
r = write_table(ws3, 4, ["Chỉ số", "Giá trị"], ct_kpi)

# By campaign
ws3.cell(row=r, column=1, value="CHI TIẾT THEO CHIẾN DỊCH CTKM THÁNG 3").font = SECTION_FONT
ct_c_headers = ["Chiến dịch", "Bắt đầu", "Kết thúc", "Số salon", "Tổng SD", "% CTKM T3", "TB SD/salon"]
ct_c_data = []
for name, info in ct_camp_sorted:
    ns = len(info['salons'])
    ct_c_data.append([
        name, info['start'], info['end'], ns, info['used'],
        round(info['used'] / ct_total * 100, 1) if ct_total else 0,
        round(info['used'] / ns, 2) if ns else 0
    ])
r = write_table(ws3, r + 1, ct_c_headers, ct_c_data)

# By salon
ws3.cell(row=r, column=1, value="CHI TIẾT THEO SALON - CTKM THÁNG 3").font = SECTION_FONT
ct_s_headers = ["Salon", "Tổng SD", "% CTKM T3", "Số chiến dịch", "TB SD/chiến dịch", "Danh sách chiến dịch"]
ct_s_data = []
for name, info in ct_salon_sorted:
    nc = len(info['campaigns'])
    ct_s_data.append([
        name, info['used'],
        round(info['used'] / ct_total * 100, 1) if ct_total else 0,
        nc,
        round(info['used'] / nc, 2) if nc else 0,
        "; ".join(sorted(info['campaigns']))[:200]
    ])
r = write_table(ws3, r + 1, ct_s_headers, ct_s_data)

# ===================== DOANH THU BRAND - CTKM THÁNG 3 =====================
BRAND_UP_FILL = PatternFill("solid", fgColor="C6EFCE")    # green - tăng
BRAND_DOWN_FILL = PatternFill("solid", fgColor="FFC7CE")   # red - giảm
BRAND_NEUTRAL_FILL = PatternFill("solid", fgColor="FFEB9C") # yellow

brand_revenue = [
    {
        'brand': 'Laborie',
        'dt_t3': 160_000_000, 'sl_t3': 503,
        'dt_t2': 176_000_000, 'sl_t2': 377,
        'dt_t1': 156_000_000, 'sl_t1': 355,
        'campaigns': [c for c, _ in ct_camp_sorted if 'laborie' in c.lower() or 'Laborie' in c],
    },
    {
        'brand': 'Máy sấy tóc',
        'dt_t3': 36_300_000, 'sl_t3': 60,
        'dt_t2': 70_000_000, 'sl_t2': 84,
        'dt_t1': 34_000_000, 'sl_t1': 55,
        'campaigns': [c for c, _ in ct_camp_sorted if 'máy sấy' in c.lower() or 'Máy sấy' in c],
    },
    {
        'brand': 'Máy Massage',
        'dt_t3': 14_000_000, 'sl_t3': 14,
        'dt_t2': 16_000_000, 'sl_t2': 21,
        'dt_t1': 41_000_000, 'sl_t1': 53,
        'campaigns': [c for c, _ in ct_camp_sorted if 'massage' in c.lower() or 'Massage' in c],
    },
]

# Calc derived metrics
for b in brand_revenue:
    b['dt_mom_t3t2'] = round((b['dt_t3'] - b['dt_t2']) / b['dt_t2'] * 100, 1) if b['dt_t2'] else 0
    b['dt_mom_t2t1'] = round((b['dt_t2'] - b['dt_t1']) / b['dt_t1'] * 100, 1) if b['dt_t1'] else 0
    b['sl_mom_t3t2'] = round((b['sl_t3'] - b['sl_t2']) / b['sl_t2'] * 100, 1) if b['sl_t2'] else 0
    b['sl_mom_t2t1'] = round((b['sl_t2'] - b['sl_t1']) / b['sl_t1'] * 100, 1) if b['sl_t1'] else 0
    b['gia_tb_t3'] = round(b['dt_t3'] / b['sl_t3']) if b['sl_t3'] else 0
    b['gia_tb_t2'] = round(b['dt_t2'] / b['sl_t2']) if b['sl_t2'] else 0
    b['gia_tb_t1'] = round(b['dt_t1'] / b['sl_t1']) if b['sl_t1'] else 0
    # CTKM usage
    b['ctkm_used'] = sum(info['used'] for c, info in ct_camp_sorted if c in b['campaigns'])
    b['ctkm_salons'] = len(set().union(*(info['salons'] for c, info in ct_camp_sorted if c in b['campaigns']))) if b['campaigns'] else 0

ws3.cell(row=r, column=1, value="DOANH THU BRAND THAM GIA CTKM THÁNG 3").font = SECTION_FONT

# Revenue table
rev_headers = [
    "Brand", "DT Tháng 1", "SL T1", "DT Tháng 2", "SL T2", "DT Tháng 3", "SL T3",
    "% DT T2→T3", "% SL T2→T3", "% DT T1→T2", "% SL T1→T2"
]
rev_data = []
for b in brand_revenue:
    rev_data.append([
        b['brand'],
        b['dt_t1'], b['sl_t1'],
        b['dt_t2'], b['sl_t2'],
        b['dt_t3'], b['sl_t3'],
        f"{b['dt_mom_t3t2']:+.1f}%", f"{b['sl_mom_t3t2']:+.1f}%",
        f"{b['dt_mom_t2t1']:+.1f}%", f"{b['sl_mom_t2t1']:+.1f}%",
    ])
r = write_table(ws3, r + 1, rev_headers, rev_data)

# Format number columns as VND
for ri in range(r - len(rev_data) - 1, r - 1):
    for ci in [2, 4, 6]:  # DT columns
        cell = ws3.cell(row=ri, column=ci)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0'

# Color the % cells
for ri in range(r - len(rev_data) - 1, r - 1):
    for ci in [8, 9, 10, 11]:
        cell = ws3.cell(row=ri, column=ci)
        if cell.value and isinstance(cell.value, str):
            val = float(cell.value.replace('%', '').replace('+', ''))
            if val > 0:
                cell.fill = BRAND_UP_FILL
            elif val < 0:
                cell.fill = BRAND_DOWN_FILL

# Average price table
ws3.cell(row=r, column=1, value="GIÁ BÁN TRUNG BÌNH THEO THÁNG").font = SECTION_FONT
price_headers = ["Brand", "Giá TB T1", "Giá TB T2", "Giá TB T3", "Xu hướng giá"]
price_data = []
for b in brand_revenue:
    if b['gia_tb_t3'] > b['gia_tb_t2']:
        trend = "↑ Tăng"
    elif b['gia_tb_t3'] < b['gia_tb_t2']:
        trend = "↓ Giảm"
    else:
        trend = "→ Ổn định"
    price_data.append([
        b['brand'], b['gia_tb_t1'], b['gia_tb_t2'], b['gia_tb_t3'], trend
    ])
r = write_table(ws3, r + 1, price_headers, price_data)

for ri in range(r - len(price_data) - 1, r - 1):
    for ci in [2, 3, 4]:
        cell = ws3.cell(row=ri, column=ci)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0'

# CTKM x Revenue mapping
ws3.cell(row=r, column=1, value="LIÊN KẾT CTKM - DOANH THU").font = SECTION_FONT
link_headers = ["Brand", "Chiến dịch CTKM T3", "Lượt SD CTKM", "Số salon CTKM", "DT T3", "SL T3", "% DT T2→T3"]
link_data = []
for b in brand_revenue:
    link_data.append([
        b['brand'],
        "; ".join(b['campaigns'])[:200] if b['campaigns'] else "Không có CT riêng",
        b['ctkm_used'],
        b['ctkm_salons'],
        b['dt_t3'],
        b['sl_t3'],
        f"{b['dt_mom_t3t2']:+.1f}%",
    ])
r = write_table(ws3, r + 1, link_headers, link_data)

for ri in range(r - len(link_data) - 1, r - 1):
    cell = ws3.cell(row=ri, column=5)
    if isinstance(cell.value, (int, float)):
        cell.number_format = '#,##0'

# ĐÁNH GIÁ TỔNG HỢP
ws3.cell(row=r, column=1, value="ĐÁNH GIÁ TỔNG HỢP").font = Font(bold=True, size=12, color="C00000")
eval_row = r + 1

evaluations = []
for b in brand_revenue:
    lines = []
    # DT trend
    if b['dt_mom_t3t2'] > 0:
        lines.append(f"DT tháng 3 TĂNG {b['dt_mom_t3t2']:+.1f}% so với T2")
    else:
        lines.append(f"DT tháng 3 GIẢM {b['dt_mom_t3t2']:.1f}% so với T2")
    # SL trend
    if b['sl_mom_t3t2'] > 0:
        lines.append(f"SL bán tháng 3 TĂNG {b['sl_mom_t3t2']:+.1f}% so với T2 ({b['sl_t2']} → {b['sl_t3']})")
    else:
        lines.append(f"SL bán tháng 3 GIẢM {b['sl_mom_t3t2']:.1f}% so với T2 ({b['sl_t2']} → {b['sl_t3']})")
    # Price insight
    if b['gia_tb_t3'] < b['gia_tb_t2']:
        lines.append(f"Giá TB giảm ({b['gia_tb_t2']:,.0f} → {b['gia_tb_t3']:,.0f}) - có thể do CTKM combo giảm giá")
    elif b['gia_tb_t3'] > b['gia_tb_t2']:
        lines.append(f"Giá TB tăng ({b['gia_tb_t2']:,.0f} → {b['gia_tb_t3']:,.0f})")
    # CTKM effectiveness
    if b['ctkm_used'] > 0:
        lines.append(f"CTKM có {b['ctkm_used']} lượt SD tại {b['ctkm_salons']} salon")
    else:
        lines.append("Không có CT CTKM Tháng 3 riêng cho brand này")
    # Overall assessment
    if b['dt_mom_t3t2'] < 0 and b['sl_mom_t3t2'] > 0:
        lines.append("→ SL tăng nhưng DT giảm: CTKM đẩy được SL nhưng giá TB giảm do combo/KM")
    elif b['dt_mom_t3t2'] < 0 and b['sl_mom_t3t2'] < 0:
        lines.append("→ Cả DT và SL đều giảm: CTKM chưa hiệu quả, cần xem lại chiến lược")
    elif b['dt_mom_t3t2'] > 0 and b['sl_mom_t3t2'] > 0:
        lines.append("→ DT và SL đều tăng: CTKM hoạt động hiệu quả")
    evaluations.append((b['brand'], lines))

for brand, lines in evaluations:
    ws3.cell(row=eval_row, column=1, value=f"● {brand}").font = SUBHEADER_FONT
    eval_row += 1
    for line in lines:
        ws3.cell(row=eval_row, column=1, value=f"   {line}")
        # Color last line (overall assessment)
        if line.startswith("→"):
            if "hiệu quả" in line and "chưa" not in line:
                ws3.cell(row=eval_row, column=1).fill = BRAND_UP_FILL
            elif "chưa hiệu quả" in line:
                ws3.cell(row=eval_row, column=1).fill = BRAND_DOWN_FILL
            else:
                ws3.cell(row=eval_row, column=1).fill = BRAND_NEUTRAL_FILL
        eval_row += 1
    eval_row += 1

r = eval_row + 1

# Raw data
ws3.cell(row=r, column=1, value="DỮ LIỆU CHI TIẾT CTKM THÁNG 3").font = SECTION_FONT
raw_data_ct = [[row['campaign'], row['start'], row['end'], row['salon'], row['used']] for row in ct_rows]
raw_data_ct.sort(key=lambda x: -x[4])
r = write_table(ws3, r + 1, raw_headers, raw_data_ct)

auto_width(ws3, 11)

# ============ SHEET 4: SO SÁNH SALON 2 DANH MỤC ============
ws4 = wb.create_sheet("Salon - Clearstock vs CTKM T3")

ws4.cell(row=1, column=1, value="PHÂN TÍCH SALON: CLEARSTOCK vs CTKM THÁNG 3").font = TITLE_FONT
ws4.merge_cells('A1:I1')

# Summary comparison
ws4.cell(row=3, column=1, value="SO SÁNH TỔNG QUAN 2 DANH MỤC").font = SECTION_FONT
comp_headers = ["Chỉ số", "Clearstock", "CTKM Tháng 3", "Tổng cộng"]
comp_data = [
    ["Tổng lượt SD", cs_total, ct_total, cs_total + ct_total],
    ["Số chiến dịch", len(cs_camp), len(ct_camp), len(cs_camp) + len(ct_camp)],
    ["Số salon", len(cs_salon), len(ct_salon), len(all_salons_both)],
    ["Số bản ghi", len(cs_rows), len(ct_rows), len(cs_rows) + len(ct_rows)],
    ["TB SD/bản ghi", round(cs_total/len(cs_rows),2) if cs_rows else 0, round(ct_total/len(ct_rows),2) if ct_rows else 0, round((cs_total+ct_total)/(len(cs_rows)+len(ct_rows)),2)],
    ["TB SD/salon", round(cs_total/len(cs_salon),2) if cs_salon else 0, round(ct_total/len(ct_salon),2) if ct_salon else 0, ""],
]
r = write_table(ws4, 4, comp_headers, comp_data)

# Salon overlap analysis
cs_only = set(cs_salon.keys()) - set(ct_salon.keys())
ct_only = set(ct_salon.keys()) - set(cs_salon.keys())
both_cats = set(cs_salon.keys()) & set(ct_salon.keys())

master_names = set(s['name'] for s in salon_master)
no_participation = master_names - all_salons_both

ws4.cell(row=r, column=1, value="PHÂN TÍCH SALON CHUNG / RIÊNG").font = SECTION_FONT
overlap_data = [
    ["Salon chỉ có Clearstock", len(cs_only), ", ".join(sorted(cs_only))[:200]],
    ["Salon chỉ có CTKM T3", len(ct_only), ", ".join(sorted(ct_only))[:200]],
    ["Salon có CẢ HAI", len(both_cats), ", ".join(sorted(both_cats))[:200]],
    ["Salon KHÔNG tham gia cả 2 (theo DS master)", len(no_participation), ", ".join(sorted(no_participation))[:200]],
    ["Tổng salon trong DS master", len(salon_master), ""],
]
r = write_table(ws4, r + 1, ["Phân loại", "Số salon", "Danh sách salon"], overlap_data)

# Full salon comparison table
ws4.cell(row=r, column=1, value="CHI TIẾT TỪNG SALON - SO SÁNH 2 DANH MỤC").font = SECTION_FONT
sb_headers = [
    "Salon", "SD Clearstock", "SD CTKM T3", "Tổng SD",
    "Số CT Clearstock", "Số CT CTKM T3",
    "% Clearstock", "% CTKM T3", "Đánh giá"
]
sb_data = []
for name, info in salon_both_sorted:
    total_s = info['total']
    cs_pct = round(info['cs_used'] / total_s * 100, 1) if total_s else 0
    ct_pct = round(info['ct_used'] / total_s * 100, 1) if total_s else 0
    # Rating
    if info['cs_used'] > 0 and info['ct_used'] > 0:
        rating = "Tham gia cả 2"
    elif info['cs_used'] > 0:
        rating = "Chỉ Clearstock"
    else:
        rating = "Chỉ CTKM T3"
    sb_data.append([
        name, info['cs_used'], info['ct_used'], total_s,
        len(info['cs_camps']), len(info['ct_camps']),
        cs_pct, ct_pct, rating
    ])
r = write_table(ws4, r + 1, sb_headers, sb_data)

# Highlight top performers
ws4.cell(row=r, column=1, value="SALON NỔI BẬT").font = SECTION_FONT
# Top 5 clearstock salons
ws4.cell(row=r + 1, column=1, value="Top 5 Salon Clearstock:").font = SUBHEADER_FONT
top5cs = cs_salon_sorted[:5]
for i, (name, info) in enumerate(top5cs):
    ws4.cell(row=r + 2 + i, column=1, value=f"  {i+1}. {name}")
    ws4.cell(row=r + 2 + i, column=2, value=f"{info['used']} lượt ({round(info['used']/cs_total*100,1)}%)")

r2 = r + 2 + len(top5cs) + 1
ws4.cell(row=r2, column=1, value="Top 5 Salon CTKM Tháng 3:").font = SUBHEADER_FONT
top5ct = ct_salon_sorted[:5]
for i, (name, info) in enumerate(top5ct):
    ws4.cell(row=r2 + 1 + i, column=1, value=f"  {i+1}. {name}")
    ws4.cell(row=r2 + 1 + i, column=2, value=f"{info['used']} lượt ({round(info['used']/ct_total*100,1)}%)")

# ===================== SALON KHÔNG THAM GIA =====================
NOT_PARTICIPATING_FILL = PatternFill("solid", fgColor="FCE4EC")

# Find salons from master list not in Clearstock or CTKM T3
master_salon_names = set(s['name'] for s in salon_master)
participating_salons = all_salons_both  # salons that have Clearstock or CTKM T3

not_participating = []
for s in salon_master:
    if s['name'] not in participating_salons:
        # Check if they participate in ANY campaign at all
        any_categories = set()
        any_used = 0
        any_camps = 0
        for row in rows:
            if row['salon'] == s['name']:
                any_categories.add(row['category'])
                any_used += row['used']
                any_camps += 1
        not_participating.append({
            **s,
            'other_categories': ", ".join(sorted(any_categories)) if any_categories else "Không có",
            'other_used': any_used,
            'other_records': any_camps,
        })

not_participating.sort(key=lambda x: x['cluster'])

r3 = r2 + 1 + len(top5ct) + 2
ws4.cell(row=r3, column=1, value="DANH SÁCH SALON KHÔNG THAM GIA CLEARSTOCK & CTKM THÁNG 3").font = SECTION_FONT
ws4.cell(row=r3, column=1).font = Font(bold=True, size=12, color="C00000")

# Summary
ws4.cell(row=r3 + 1, column=1, value=f"Tổng: {len(not_participating)} salon không tham gia Clearstock hoặc CTKM Tháng 3 (trên tổng {len(salon_master)} salon)")

np_headers = [
    "ID", "Tên Salon", "Supervisor", "Phân cụm", "Loại (online/offline)",
    "Danh mục khác đang tham gia", "Lượt SD danh mục khác", "Số bản ghi khác"
]
np_data = []
for s in not_participating:
    np_data.append([
        s['id'], s['name'], s['supervisor'], s['cluster'], s['type'],
        s['other_categories'], s['other_used'], s['other_records']
    ])
r4 = write_table(ws4, r3 + 3, np_headers, np_data)

# Highlight rows with no participation at all
for ri in range(r3 + 4, r3 + 4 + len(np_data)):
    cell_val = ws4.cell(row=ri, column=6).value
    if cell_val == "Không có":
        for ci in range(1, len(np_headers) + 1):
            ws4.cell(row=ri, column=ci).fill = NOT_PARTICIPATING_FILL

# Summary by cluster
ws4.cell(row=r4, column=1, value="THỐNG KÊ SALON KHÔNG THAM GIA THEO CỤM").font = SECTION_FONT
cluster_stats = defaultdict(lambda: {'total_master': 0, 'not_participating': 0, 'names': []})
for s in salon_master:
    cluster_stats[s['cluster']]['total_master'] += 1
for s in not_participating:
    cluster_stats[s['cluster']]['not_participating'] += 1
    cluster_stats[s['cluster']]['names'].append(s['name'])

cl_headers = ["Phân cụm", "Tổng salon trong cụm", "Không tham gia", "% Không tham gia", "Danh sách salon"]
cl_data = []
for cluster in sorted(cluster_stats.keys()):
    info = cluster_stats[cluster]
    pct = round(info['not_participating'] / info['total_master'] * 100, 1) if info['total_master'] else 0
    cl_data.append([
        cluster, info['total_master'], info['not_participating'], pct,
        ", ".join(sorted(info['names']))[:200]
    ])
cl_data.sort(key=lambda x: -x[3])  # sort by % not participating
r5 = write_table(ws4, r4 + 1, cl_headers, cl_data)

# Summary by supervisor
ws4.cell(row=r5, column=1, value="THỐNG KÊ SALON KHÔNG THAM GIA THEO SUPERVISOR").font = SECTION_FONT
sup_stats = defaultdict(lambda: {'total_master': 0, 'not_participating': 0, 'names': []})
for s in salon_master:
    sup_stats[s['supervisor']]['total_master'] += 1
for s in not_participating:
    sup_stats[s['supervisor']]['not_participating'] += 1
    sup_stats[s['supervisor']]['names'].append(s['name'])

sp_headers = ["Supervisor", "Tổng salon quản lý", "Không tham gia", "% Không tham gia", "Danh sách salon"]
sp_data = []
for sup in sorted(sup_stats.keys()):
    info = sup_stats[sup]
    pct = round(info['not_participating'] / info['total_master'] * 100, 1) if info['total_master'] else 0
    sp_data.append([
        sup, info['total_master'], info['not_participating'], pct,
        ", ".join(sorted(info['names']))[:200]
    ])
sp_data.sort(key=lambda x: -x[3])
r6 = write_table(ws4, r5 + 1, sp_headers, sp_data)

auto_width(ws4, 9)

# ============ SHEET 5: DỮ LIỆU GỐC ============
ws5 = wb.create_sheet("Dữ liệu gốc")
ws5.cell(row=1, column=1, value="TOÀN BỘ DỮ LIỆU CTKM THÁNG 3").font = TITLE_FONT
all_headers = ["Chiến dịch", "Bắt đầu", "Kết thúc", "Salon", "Đã sử dụng", "Danh mục"]
for i, h in enumerate(all_headers, 1):
    ws5.cell(row=3, column=i, value=h)
style_header(ws5, 3, 6)
for ri, row in enumerate(rows, 4):
    ws5.cell(row=ri, column=1, value=row['campaign'])
    ws5.cell(row=ri, column=2, value=row['start'])
    ws5.cell(row=ri, column=3, value=row['end'])
    ws5.cell(row=ri, column=4, value=row['salon'])
    ws5.cell(row=ri, column=5, value=row['used'])
    ws5.cell(row=ri, column=6, value=row['category'])
style_data_rows(ws5, 4, 3 + len(rows), 6)
auto_width(ws5, 6)

# ============ SHEET 6: ĐÁNH GIÁ CTKM THEO 3 NHÓM ============
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")

ws6 = wb.create_sheet("Đánh giá CTKM - 3 Nhóm")

ws6.cell(row=1, column=1, value="ĐÁNH GIÁ CTKM THEO 3 NHÓM SẢN PHẨM").font = TITLE_FONT
ws6.merge_cells('A1:H1')

# --- Bảng ngưỡng rule ---
ws6.cell(row=3, column=1, value="BẢNG NGƯỠNG RULE THEO NHÓM").font = SECTION_FONT
rule_ref_headers = ["Tiêu chí", "A: SP Mới về", "B: SP Thường", "C: Tồn date dài"]
rule_ref_data = [
    ["Margin Rate tối thiểu", ">= 30%", ">= 40%", ">= 15%"],
    ["Giảm giá tối đa", "<= 20%", "<= 30%", "<= 50%"],
    ["Độ phủ salon", ">= 50%", "Chọn lọc", ">= 80%"],
    ["KPI trọng tâm", "ROI >= 80%", "% KH mới >= 70%", "Độ phủ salon >= 80%"],
    ["KPI phụ", "Độ phủ salon >= 50%", "CAC <= 2x Margin", "Margin Rate >= 15%"],
]
r6 = write_table(ws6, 4, rule_ref_headers, rule_ref_data)

# --- Bảng COGS giải thích ---
ws6.cell(row=r6, column=1, value="CÔNG THỨC COGS & MARGIN RATE").font = SECTION_FONT
formula_headers = ["Công thức", "Giải thích", "Ví dụ"]
formula_data = [
    ["COGS = Giá nhập x (1 + % Vận hành)", "Giá vốn = Giá nhập + Vận chuyển + Hao hụt + Hoa hồng", "100,000 x 1.3 = 130,000đ"],
    ["Margin = Giá bán KM - COGS", "Lợi nhuận mỗi lượt SD", "280,000 - 130,000 = 150,000đ"],
    ["Margin Rate = Margin / Giá bán KM x 100", "% lợi nhuận trên giá bán KM", "150,000 / 280,000 = 53.6%"],
    ["Chi phí KM = Giá gốc - Giá bán KM", "Số tiền giảm giá mỗi lượt", "400,000 - 280,000 = 120,000đ"],
    ["ROI = Tổng Margin / Tổng Chi phí KM x 100", "1đ giảm giá sinh bao nhiêu đ lời", "65.4M / 60.4M = 108.3%"],
    ["CAC = Tổng Chi phí KM / Số KH mới", "Chi phí để có 1 KH mới", "Chỉ Nhóm B"],
    ["Vòng quay vốn = DT bán KM / (COGS x Tồn kho)", "Số lần thu hồi vốn trên tồn kho", "Chỉ Nhóm C"],
]
r6 = write_table(ws6, r6 + 1, formula_headers, formula_data)

# --- Bảng tổng hợp đánh giá ---
ws6.cell(row=r6, column=1, value="TỔNG HỢP ĐÁNH GIÁ").font = SECTION_FONT
sum_headers = [
    "Sản phẩm", "Nhóm", "Giá gốc", "COGS", "Giá KM",
    "Margin Rate", "Tổng lượt SD", "Số salon", "Độ phủ",
    "ROI", "Pass/Total", "Kết quả"
]
sum_data = []
for ev in product_eval:
    sum_data.append([
        ev['name'], f"{ev['group']}: {ev['group_name']}",
        ev['gia_goc'], ev['cogs'], ev['gia_km'],
        f"{ev['margin_rate']}%", ev['total_sd'], ev['num_salons'],
        f"{ev['do_phu']}%", f"{ev['roi']}%",
        f"{ev['pass_count']}/{ev['total_checks']}", ev['verdict'],
    ])
r6 = write_table(ws6, r6 + 1, sum_headers, sum_data)

# Tô màu kết quả + format số
for ri in range(r6 - len(sum_data) - 1, r6 - 1):
    # Verdict color
    cell_v = ws6.cell(row=ri, column=12)
    if cell_v.value == 'ĐẠT':
        cell_v.fill = PASS_FILL
    elif cell_v.value == 'KHÔNG ĐẠT':
        cell_v.fill = FAIL_FILL
    else:
        cell_v.fill = WARN_FILL
    # Format VND columns
    for ci in [3, 4, 5]:
        cell_n = ws6.cell(row=ri, column=ci)
        if isinstance(cell_n.value, (int, float)):
            cell_n.number_format = '#,##0'

# --- Chi tiết từng SP ---
for ev in product_eval:
    ws6.cell(row=r6, column=1,
             value=f"CHI TIẾT: {ev['name']} (Nhóm {ev['group']}: {ev['group_name']})").font = SECTION_FONT

    # Thông tin cơ bản
    info_data = [
        ["Giá gốc", f"{ev['gia_goc']:,}đ"],
        ["COGS (Giá vốn)", f"{ev['cogs']:,}đ"],
        ["Giá bán KM", f"{ev['gia_km']:,}đ"],
        ["Margin/lượt", f"{ev['margin']:,}đ"],
        ["Mức giảm giá", f"{ev['discount']:,}đ ({ev['discount_pct']}%)"],
        ["Margin Rate", f"{ev['margin_rate']}%"],
        ["Tổng lượt SD", ev['total_sd']],
        ["Số salon tham gia", ev['num_salons']],
        ["Độ phủ salon", f"{ev['do_phu']}%"],
        ["Thời gian KM", f"{ev['duration']} ngày"],
        ["Tổng margin", f"{ev['total_margin']:,}đ"],
        ["Tổng chi phí KM", f"{ev['total_cost_km']:,}đ"],
        ["ROI trên margin", f"{ev['roi']}%"],
    ]

    # Thêm KPI đặc thù theo nhóm
    if ev['group'] == 'B':
        info_data.append(["% KH mới (ước tính)", f"{ev.get('pct_kh_moi', 0)}%"])
        info_data.append(["Số KH mới", ev.get('so_kh_moi', 0)])
        info_data.append(["CAC (Chi phí/KH mới)", f"{ev.get('cac', 0):,}đ"])
        info_data.append(["CAC giới hạn (2x Margin)", f"{ev.get('cac_limit', 0):,}đ"])
        info_data.append(["Payback (số lần mua hoà vốn)", f"{ev.get('payback', 0)} lần"])
    elif ev['group'] == 'C':
        info_data.append(["Tồn kho ban đầu", ev.get('ton_kho', 0)])
        info_data.append(["Tồn kho còn lại", ev.get('ton_con', 0)])
        info_data.append(["% Đã xả", f"{ev.get('pct_xa', 0)}%"])
        info_data.append(["Vòng quay vốn", ev.get('vong_quay', 0)])
        info_data.append(["Burn Rate", f"{ev.get('burn_rate_week', 0)} lượt/tuần"])
        if 'days_left' in ev:
            info_data.append(["Số ngày còn date", ev['days_left']])
        if 'weeks_need' in ev:
            info_data.append(["Tuần cần để xả hết", ev['weeks_need']])
            info_data.append(["Tuần còn lại", ev['weeks_avail']])

    r6 = write_table(ws6, r6 + 1, ["Chỉ số", "Giá trị"], info_data)

    # Bảng kiểm tra rule
    ws6.cell(row=r6, column=1, value="KIỂM TRA RULE").font = SUBHEADER_FONT
    ck_headers = ["Rule", "Thực tế", "Ngưỡng", "Kết quả"]
    ck_data = []
    for ck in ev['checks']:
        ck_data.append([ck['rule'], ck['actual'], ck['target'], 'PASS' if ck['ok'] else 'FAIL'])
    r6 = write_table(ws6, r6 + 1, ck_headers, ck_data)

    # Tô màu PASS/FAIL
    for ri in range(r6 - len(ck_data) - 1, r6 - 1):
        cell_ck = ws6.cell(row=ri, column=4)
        if cell_ck.value == 'PASS':
            cell_ck.fill = PASS_FILL
        elif cell_ck.value == 'FAIL':
            cell_ck.fill = FAIL_FILL

    # Kết quả + Đề xuất
    verdict_text = f"KẾT QUẢ: {ev['verdict']} ({ev['pass_count']}/{ev['total_checks']} rules passed)"
    ws6.cell(row=r6, column=1, value=verdict_text).font = Font(bold=True, size=11)
    if ev['verdict'] == 'ĐẠT':
        ws6.cell(row=r6, column=1).fill = PASS_FILL
    elif ev['verdict'] == 'KHÔNG ĐẠT':
        ws6.cell(row=r6, column=1).fill = FAIL_FILL
    else:
        ws6.cell(row=r6, column=1).fill = WARN_FILL
    r6 += 1

    ws6.cell(row=r6, column=1, value="ĐỀ XUẤT:").font = SUBHEADER_FONT
    r6 += 1
    for rec in ev['recs']:
        ws6.cell(row=r6, column=1, value=f"  - {rec}")
        r6 += 1

    # Danh sách campaign
    ws6.cell(row=r6, column=1, value="Chiến dịch: " + "; ".join(ev['campaigns'])[:300])
    r6 += 3

auto_width(ws6, 12)

# ============ SHEET 7: VER 2 — BỘ RULE ĐÁNH GIÁ CTKM ĐƠN GIẢN ============
ws7 = wb.create_sheet("Ver 2 - Rule đánh giá CTKM")

ws7.cell(row=1, column=1, value="VER 2 — BỘ RULE ĐÁNH GIÁ CTKM CHO SP ĐANG BÁN TẠI SALON").font = TITLE_FONT
ws7.merge_cells('A1:F1')
ws7.cell(row=2, column=1, value="(Không áp dụng cho SP cận date / Clearstock — có rule riêng)").font = Font(italic=True, color="666666")

# === TẦNG 1 ===
r7 = 4
ws7.cell(row=r7, column=1, value="TẦNG 1 — LUÔN KIỂM TRA (không cần kỳ trước)").font = SECTION_FONT
t1_headers = ["#", "Rule", "Công thức", "Ngưỡng", "Ý nghĩa"]
t1_data = [
    [1, "Margin Rate", "(Giá bán KM - COGS) / Giá bán KM x 100", ">= 20%", "Mỗi lượt bán KM vẫn có lời"],
    [2, "Mức giảm giá", "(Giá gốc - Giá bán KM) / Giá gốc x 100", "<= 40%", "Không phá giá sản phẩm"],
]
r7 = write_table(ws7, r7 + 1, t1_headers, t1_data)

# === TẦNG 2 ===
ws7.cell(row=r7, column=1, value="TẦNG 2 — SO SÁNH LNG VỚI KỲ TRƯỚC (khi có data kỳ trước)").font = SECTION_FONT
t2_headers = ["#", "Rule", "Công thức", "Ngưỡng", "Ý nghĩa"]
t2_data = [
    [3, "Margin Rate không giảm quá sâu", "MR kỳ KM - MR kỳ trước", ">= -10 điểm %", "KM không được ăn quá 10 điểm margin so với kỳ trước"],
    [4, "Margin/lượt không giảm quá sâu", "Margin kỳ KM / Margin kỳ trước x 100", ">= 70%", "Lời mỗi lượt bán không giảm quá 30% so với kỳ trước"],
    [5, "Giá TB không giảm quá sâu", "Giá TB kỳ KM / Giá TB kỳ trước x 100", ">= 75%", "Combo/KM không kéo giá TB xuống quá 25%"],
]
r7 = write_table(ws7, r7 + 1, t2_headers, t2_data)

# === TẦNG 3 ===
ws7.cell(row=r7, column=1, value="TẦNG 3 — TUỲ CHỌN (bật khi kỳ trước tương đương, tắt khi Tết/cao điểm)").font = SECTION_FONT
t3_headers = ["#", "Rule", "Công thức", "Ngưỡng", "Khi nào bật", "Khi nào tắt"]
t3_data = [
    [6, "DT tăng trưởng", "DT kỳ KM / DT kỳ trước - 1", ">= 0% (không giảm)", "Kỳ trước bình thường, cùng mùa", "Kỳ trước là Tết / cao điểm / SP mới"],
    [7, "SL tăng trưởng", "SL kỳ KM / SL kỳ trước - 1", ">= 0% (không giảm)", "Kỳ trước bình thường, cùng mùa", "Kỳ trước là Tết / cao điểm / SP mới"],
]
r7 = write_table(ws7, r7 + 1, t3_headers, t3_data)

# === BẢNG CÔNG THỨC ===
ws7.cell(row=r7, column=1, value="CÔNG THỨC TÍNH").font = SECTION_FONT
f2_headers = ["Chỉ số", "Công thức", "Ví dụ (Laborie)"]
f2_data = [
    ["COGS (Giá vốn)", "Giá nhập x (1 + % Vận hành)", "100,000 x 1.3 = 130,000đ"],
    ["Giá bán KM", "Giá gốc - Mức giảm", "400,000 - 120,000 = 280,000đ"],
    ["Margin/lượt", "Giá bán KM - COGS", "280,000 - 130,000 = 150,000đ"],
    ["Margin Rate (%)", "(Giá bán KM - COGS) / Giá bán KM x 100", "(280,000 - 130,000) / 280,000 = 53.6%"],
    ["Mức giảm giá (%)", "(Giá gốc - Giá bán KM) / Giá gốc x 100", "(400,000 - 280,000) / 400,000 = 30.0%"],
    ["Biến động MR", "MR kỳ KM - MR kỳ trước", "53.6% - 67.5% = -13.9 điểm %"],
    ["Tỷ lệ Margin/lượt", "Margin kỳ KM / Margin kỳ trước x 100", "150,000 / 270,000 = 55.6%"],
    ["Tỷ lệ Giá TB", "Giá TB kỳ KM / Giá TB kỳ trước x 100", "318,000 / 467,000 = 68.1%"],
    ["DT tăng trưởng", "DT kỳ KM / DT kỳ trước - 1", "160M / 176M - 1 = -9.1%"],
    ["SL tăng trưởng", "SL kỳ KM / SL kỳ trước - 1", "503 / 377 - 1 = +33.4%"],
]
r7 = write_table(ws7, r7 + 1, f2_headers, f2_data)

# === BẢNG NGƯỠNG ĐÁNH GIÁ ===
ws7.cell(row=r7, column=1, value="BẢNG NGƯỠNG ĐÁNH GIÁ TỔNG HỢP").font = SECTION_FONT
ng_headers = ["Kết quả", "Điều kiện", "Hành động"]
ng_data = [
    ["ĐẠT", "Tất cả rule PASS", "Tiếp tục triển khai"],
    ["CẦN CẢI THIỆN", ">= 50% rule PASS", "Điều chỉnh mức KM hoặc thu hẹp phạm vi"],
    ["KHÔNG ĐẠT", "< 50% rule PASS", "Dừng CTKM, xem lại chiến lược giá"],
]
r7 = write_table(ws7, r7 + 1, ng_headers, ng_data)

# Tô màu kết quả
for ri in range(r7 - len(ng_data) - 1, r7 - 1):
    cell_ng = ws7.cell(row=ri, column=1)
    if cell_ng.value == 'ĐẠT':
        cell_ng.fill = PASS_FILL
    elif cell_ng.value == 'KHÔNG ĐẠT':
        cell_ng.fill = FAIL_FILL
    elif cell_ng.value == 'CẦN CẢI THIỆN':
        cell_ng.fill = WARN_FILL

# === VÍ DỤ MINH HOẠ ===
ws7.cell(row=r7, column=1, value="VÍ DỤ MINH HOẠ CÁCH ÁP DỤNG").font = SECTION_FONT
vd_headers = ["SP", "MR kỳ KM", "Giảm giá", "Biến động MR", "Margin/lượt", "Giá TB", "DT MoM", "SL MoM", "Tầng 3?", "Kết quả"]
vd_data = [
    ["Laborie (có kỳ trước, bình thường)", "53.6%", "30%", "-13.9 đ.%", "55.6%", "68.1%", "-9.1%", "+33.4%", "Bật", "4/7 → Cần cải thiện"],
    ["Máy sấy (SP mới, không có kỳ trước)", "42.1%", "13.6%", "—", "—", "—", "—", "—", "Tắt", "2/2 → Đạt"],
    ["SP X (có kỳ trước là Tết)", "35%", "25%", "-5 đ.%", "80%", "85%", "—", "—", "Tắt", "5/5 → Đạt"],
    ["SP Y (margin quá thấp)", "12%", "55%", "-20 đ.%", "40%", "60%", "-15%", "-5%", "Bật", "1/7 → Không đạt"],
]
r7 = write_table(ws7, r7 + 1, vd_headers, vd_data)

# === HƯỚNG DẪN CONFIG ===
ws7.cell(row=r7, column=1, value="HƯỚNG DẪN CẤU HÌNH CHO TỪNG SP").font = SECTION_FONT
hd_headers = ["Trường", "Bắt buộc?", "Giải thích"]
hd_data = [
    ["keyword", "Có", "Từ khoá tìm trong tên campaign (vd: 'laborie')"],
    ["name", "Có", "Tên SP hiển thị trong báo cáo"],
    ["gia_goc", "Có", "Giá bán niêm yết (VND)"],
    ["cogs", "Có", "Giá vốn = Giá nhập x (1 + % vận hành)"],
    ["gia_km", "Có", "Giá bán sau KM (VND)"],
    ["ky_truoc.ten_ky", "Không", "Tên kỳ so sánh (vd: 'Tháng 2')"],
    ["ky_truoc.margin_rate", "Không", "Margin Rate kỳ trước (%)"],
    ["ky_truoc.margin_per_unit", "Không", "Margin/lượt kỳ trước (VND)"],
    ["ky_truoc.gia_ban_tb", "Không", "Giá bán TB kỳ trước (VND)"],
    ["ky_truoc.dt", "Không", "Doanh thu kỳ trước (VND) — chỉ khi bật Tầng 3"],
    ["ky_truoc.sl", "Không", "Số lượng bán kỳ trước — chỉ khi bật Tầng 3"],
    ["ky_truoc.so_sanh_dt_sl", "Không", "True = bật Tầng 3, False = tắt (mặc định False)"],
]
r7 = write_table(ws7, r7 + 1, hd_headers, hd_data)

auto_width(ws7, 10)

# Save
output = r"d:\Downloads\CTKM t4\Bao_cao_CTKM_Thang3_30Shine.xlsx"
wb.save(output)
print(f"Report saved to: {output}")
print(f"Total rows: {len(rows)}, Campaigns: {len(unique_campaigns)}, Salons: {len(unique_salons)}")
print(f"Clearstock: {len(cs_rows)} rows, {cs_total} uses, {len(cs_camp)} campaigns, {len(cs_salon)} salons")
print(f"CTKM T3: {len(ct_rows)} rows, {ct_total} uses, {len(ct_camp)} campaigns, {len(ct_salon)} salons")
print(f"Salons in both: {len(both_cats)}, CS only: {len(cs_only)}, CT only: {len(ct_only)}")
